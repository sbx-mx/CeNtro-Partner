from __future__ import annotations

import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from validate_workbook import audit_workbook  # noqa: E402

SOURCE = ROOT / "docs" / "data" / "Base_CeNtro Partner.xlsx"


class DataPipelineTests(unittest.TestCase):
    def test_current_workbook_has_july_sr_and_no_blocking_issues(self) -> None:
        report = audit_workbook(SOURCE)
        self.assertEqual(report["issueCount"], 0)
        self.assertEqual(report["periodCoverage"]["latestPeriod"], "jul")
        self.assertNotIn("SR%", report["periodCoverage"]["missingHeaders"])

    def test_missing_latest_month_is_blocking(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            candidate = Path(temporary) / "missing-july.xlsx"
            shutil.copy2(SOURCE, candidate)
            workbook = load_workbook(candidate)
            sheet = workbook["SR%"]
            july_column = next(cell.column for cell in sheet[1] if str(cell.value).strip().casefold() == "jul")
            sheet.delete_cols(july_column)
            workbook.save(candidate)
            report = audit_workbook(candidate)
            self.assertGreater(report["issueCount"], 0)
            self.assertEqual(report["periodCoverage"]["missingHeaders"]["SR%"], ["jul"])

    def test_future_month_is_detected_without_code_changes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            candidate = Path(temporary) / "partial-august.xlsx"
            shutil.copy2(SOURCE, candidate)
            workbook = load_workbook(candidate)
            sheet = workbook["ADT_AA"]
            sheet.cell(1, sheet.max_column + 1, "ago")
            sheet.cell(2, sheet.max_column, 1.01)
            workbook.save(candidate)
            report = audit_workbook(candidate)
            self.assertEqual(report["periodCoverage"]["latestPeriod"], "ago")
            self.assertIn("SR%", report["periodCoverage"]["missingHeaders"])

    def test_formula_in_data_is_blocking(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            candidate = Path(temporary) / "formula.xlsx"
            shutil.copy2(SOURCE, candidate)
            workbook = load_workbook(candidate)
            workbook["SR%"]["B2"] = "=1+1"
            workbook.save(candidate)
            report = audit_workbook(candidate)
            self.assertIn({"sheet": "SR%", "cell": "B2"}, report["security"]["formulaCells"])
            self.assertGreater(report["issueCount"], 0)

    def test_sync_updates_both_destinations_with_identical_hashes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            project = Path(temporary)
            completed = subprocess.run(
                [sys.executable, str(SCRIPTS / "sync_workbook.py"), str(SOURCE), "--project", str(project)],
                check=True,
                capture_output=True,
                text=True,
            )
            result = json.loads(completed.stdout)
            public = project / "public" / "data" / SOURCE.name
            docs = project / "docs" / "data" / SOURCE.name
            self.assertEqual(public.read_bytes(), docs.read_bytes())
            self.assertEqual(result["latestPeriod"], "jul")


if __name__ == "__main__":
    unittest.main()
