#!/usr/bin/env python3
"""Valida Base_CeNtro Partner.xlsx por pestaña, encabezado y celda."""

from __future__ import annotations

import argparse
import hashlib
import json
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MONTHS = ("ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic")
PERCENT_SHEETS = {"BB", "BT", "SS"}
REQUIRED_SHEETS = {
    "Directorio", "Instrucciones", "ADT_AA", "VMT_AA%", "V_ppto", "V_AT_AA%", "vCOGS",
    "SegundasCx", "NPS", "Conexion", "Desempeño", "Bebida", "SR% ", "Rotacion", "Bajas<90",
    "Estabilidad 12M", "BB", "BT", "SS",
}
DIRECTORY_HEADERS = ("CeCo", "Tienda", "Región", "DM", "Fecha Apertura", "Tipo Tienda")
PERIOD_SHEETS = REQUIRED_SHEETS.difference({"Directorio", "Instrucciones"})
MAX_ARCHIVE_ENTRIES = 10_000
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200


def normalized(value: Any) -> str:
    return str(value or "").strip().casefold()


def clean_ceco(value: Any) -> str:
    text = str(value or "").strip().removesuffix(".0")
    digits = "".join(character for character in text if character.isdigit())
    return digits.zfill(5) if digits else ""


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def valid_opening_date(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    text = str(value or "").strip()
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            datetime.strptime(text, pattern)
            return True
        except ValueError:
            continue
    return False


def validate_xlsx_container(source: Path) -> None:
    """Rechaza archivos XLSX corruptos o con expansión desproporcionada."""
    if source.suffix.casefold() != ".xlsx" or not source.is_file():
        raise ValueError("El origen debe ser un archivo .xlsx existente.")
    if not zipfile.is_zipfile(source):
        raise ValueError("El archivo no es un contenedor XLSX válido.")
    with zipfile.ZipFile(source) as archive:
        entries = archive.infolist()
        if len(entries) > MAX_ARCHIVE_ENTRIES:
            raise ValueError("El XLSX contiene demasiados archivos internos.")
        total_size = sum(entry.file_size for entry in entries)
        if total_size > MAX_UNCOMPRESSED_BYTES:
            raise ValueError("El XLSX supera el límite seguro descomprimido de 100 MiB.")
        for entry in entries:
            if entry.filename.startswith(("/", "\\")) or ".." in Path(entry.filename).parts:
                raise ValueError(f"Ruta interna insegura en XLSX: {entry.filename}")
            compressed = max(entry.compress_size, 1)
            if entry.file_size > 1024 * 1024 and entry.file_size / compressed > MAX_COMPRESSION_RATIO:
                raise ValueError(f"Compresión sospechosa en XLSX: {entry.filename}")


def audit_workbook(source: Path) -> dict[str, Any]:
    validate_xlsx_container(source)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    workbook = load_workbook(source, data_only=False, read_only=True, keep_links=False)
    available_sheets = {normalized(name) for name in workbook.sheetnames}
    normalized_required_sheets = {normalized(name) for name in REQUIRED_SHEETS}
    normalized_period_sheets = {normalized(name) for name in PERIOD_SHEETS}
    missing_sheets = sorted(sheet for sheet in REQUIRED_SHEETS if normalized(sheet) not in available_sheets)
    sheets: list[dict[str, Any]] = []
    totals = Counter(rows=0, cells=0, blanks=0, formulas=0, numbers=0, text=0, dates=0)
    directory_metadata = {
        "records": 0,
        "regions": {},
        "districts": {},
        "storeTypes": {},
        "missingNames": [],
        "missingRegions": [],
        "missingDistricts": [],
        "missingOpeningDates": [],
        "invalidOpeningDates": [],
        "missingStoreTypes": [],
    }
    period_headers: dict[str, set[str]] = {}
    period_value_counts: dict[str, dict[str, int]] = {}
    formula_cells: list[dict[str, str]] = []
    duplicate_headers: dict[str, list[str]] = {}

    for sheet in workbook.worksheets:
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_lookup = {normalized(header): index for index, header in enumerate(headers)}
        normalized_headers = [normalized(header) for header in headers if normalized(header)]
        repeated_headers = sorted(header for header, count in Counter(normalized_headers).items() if count > 1)
        if repeated_headers:
            duplicate_headers[sheet.title] = repeated_headers
        ceco_index = header_lookup.get("ceco")
        cecos: list[str] = []
        invalid_cecos: list[str] = []
        invalid_percentages: list[dict[str, Any]] = []
        counters = Counter(rows=max(sheet.max_row - 1, 0), cells=sheet.max_row * sheet.max_column)
        monthly_counts = Counter({month: 0 for month in MONTHS})

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if ceco_index is not None:
                raw_ceco = row[ceco_index].value
                ceco = clean_ceco(raw_ceco)
                if ceco:
                    cecos.append(ceco)
                    if len(ceco) != 5:
                        invalid_cecos.append(str(raw_ceco))
            for column_number, cell in enumerate(row, start=1):
                value = cell.value
                if value in (None, ""):
                    counters["blanks"] += 1
                elif cell.data_type == "f":
                    counters["formulas"] += 1
                    formula_cells.append({"sheet": sheet.title, "cell": cell.coordinate})
                elif isinstance(value, bool):
                    counters["booleans"] += 1
                elif isinstance(value, (int, float)):
                    counters["numbers"] += 1
                elif isinstance(value, (datetime, date)):
                    counters["dates"] += 1
                else:
                    counters["text"] += 1

                header = normalized(headers[column_number - 1]) if column_number <= len(headers) else ""
                if header in MONTHS and value not in (None, ""):
                    monthly_counts[header] += 1
                if sheet.title in PERCENT_SHEETS and header in MONTHS and value not in (None, "", "N/A", "NA"):
                    if not isinstance(value, (int, float)) or isinstance(value, bool) or not 0 <= float(value) <= 1:
                        invalid_percentages.append({"cell": cell.coordinate, "value": json_value(value)})

        duplicate_cecos = sorted(ceco for ceco, count in Counter(cecos).items() if count > 1)
        missing_headers = []
        if sheet.title == "Directorio":
            missing_headers = [header for header in DIRECTORY_HEADERS if normalized(header) not in header_lookup]
            opening_index = header_lookup.get(normalized("Fecha Apertura"))
            type_index = header_lookup.get(normalized("Tipo Tienda"))
            store_index = header_lookup.get(normalized("Tienda"))
            region_index = header_lookup.get(normalized("Región"))
            district_index = header_lookup.get(normalized("DM"))
            store_types: Counter[str] = Counter()
            regions: Counter[str] = Counter()
            districts: Counter[str] = Counter()
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                ceco = clean_ceco(row[ceco_index]) if ceco_index is not None else ""
                if not ceco:
                    continue
                directory_metadata["records"] += 1
                opening = row[opening_index] if opening_index is not None else None
                store_type = str(row[type_index] or "").strip() if type_index is not None else ""
                store_name = str(row[store_index] or "").strip() if store_index is not None else ""
                region = str(row[region_index] or "").strip() if region_index is not None else ""
                district = str(row[district_index] or "").strip() if district_index is not None else ""
                if not store_name:
                    directory_metadata["missingNames"].append({"row": row_number, "ceco": ceco})
                if not region:
                    directory_metadata["missingRegions"].append({"row": row_number, "ceco": ceco})
                else:
                    regions[region] += 1
                if not district:
                    directory_metadata["missingDistricts"].append({"row": row_number, "ceco": ceco})
                else:
                    districts[district] += 1
                if not opening:
                    directory_metadata["missingOpeningDates"].append({"row": row_number, "ceco": ceco})
                elif not valid_opening_date(opening):
                    directory_metadata["invalidOpeningDates"].append({"row": row_number, "ceco": ceco, "value": json_value(opening)})
                if not store_type:
                    directory_metadata["missingStoreTypes"].append({"row": row_number, "ceco": ceco})
                if store_type:
                    store_types[store_type] += 1
            directory_metadata["storeTypes"] = dict(sorted(store_types.items()))
            directory_metadata["regions"] = dict(sorted(regions.items()))
            directory_metadata["districts"] = dict(sorted(districts.items()))
        elif sheet.title == "Instrucciones":
            expected = ("Pestaña", "Area", "Ponderacion", "Logica Selección Mes Multiple", "Logica YTD")
            missing_headers = [header for header in expected if normalized(header) not in header_lookup]
        elif normalized(sheet.title) in normalized_required_sheets:
            missing_headers = ["CeCo"] if "ceco" not in header_lookup else []

        if normalized(sheet.title) in normalized_period_sheets:
            period_headers[sheet.title] = {month for month in MONTHS if month in header_lookup}
            period_value_counts[sheet.title] = {month: monthly_counts[month] for month in MONTHS if month in header_lookup}

        totals.update(counters)
        sheets.append({
            "sheet": sheet.title,
            "rows": counters["rows"],
            "columns": sheet.max_column,
            "headers": [json_value(header) for header in headers],
            "missingHeaders": missing_headers,
            "validCeCos": len(set(cecos)),
            "duplicateCeCos": duplicate_cecos,
            "invalidCeCos": sorted(set(invalid_cecos)),
            "invalidPercentages": invalid_percentages,
            "types": {key: counters[key] for key in ("numbers", "text", "dates", "booleans", "formulas", "blanks")},
        })

    observed_months = [
        month for month in MONTHS
        if any(month in headers and period_value_counts[sheet].get(month, 0) > 0 for sheet, headers in period_headers.items())
    ]
    latest_period = observed_months[-1] if observed_months else None
    required_through_latest = set(MONTHS[: MONTHS.index(latest_period) + 1]) if latest_period else set()
    missing_period_headers = {
        sheet: sorted(required_through_latest.difference(headers), key=MONTHS.index)
        for sheet, headers in period_headers.items()
        if required_through_latest.difference(headers)
    }
    empty_latest_period_sheets = sorted(
        sheet for sheet, counts in period_value_counts.items()
        if latest_period and latest_period in period_headers[sheet] and counts.get(latest_period, 0) == 0
    )
    period_coverage_issues = sum(len(values) for values in missing_period_headers.values()) + len(empty_latest_period_sheets)

    blocking_issues = len(missing_sheets) + sum(
        len(item["missingHeaders"]) + len(item["invalidCeCos"]) + len(item["invalidPercentages"])
        + (len(item["duplicateCeCos"]) if item["sheet"] == "Directorio" else 0)
        for item in sheets
    ) + sum(len(directory_metadata[key]) for key in (
        "missingNames", "missingRegions", "missingDistricts", "missingOpeningDates",
        "invalidOpeningDates", "missingStoreTypes",
    )) + len(formula_cells) + sum(len(values) for values in duplicate_headers.values()) + period_coverage_issues
    duplicate_indicator_rows = {
        item["sheet"]: item["duplicateCeCos"] for item in sheets
        if item["sheet"] != "Directorio" and item["duplicateCeCos"]
    }
    warnings = sum(len(values) for values in duplicate_indicator_rows.values())
    return {
        "schemaVersion": 2,
        "source": source.name,
        "sha256": digest,
        "workbook": {"sheetCount": len(workbook.sheetnames), "missingSheets": missing_sheets},
        "totals": dict(totals),
        "issueCount": blocking_issues,
        "warningCount": warnings,
        "periodCoverage": {
            "latestPeriod": latest_period,
            "observedMonths": observed_months,
            "missingHeaders": missing_period_headers,
            "emptyLatestPeriodSheets": empty_latest_period_sheets,
            "valueCounts": period_value_counts,
        },
        "security": {"formulaCells": formula_cells, "duplicateHeaders": duplicate_headers},
        "directory": directory_metadata,
        "duplicateIndicatorRows": duplicate_indicator_rows,
        "sheets": sheets,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    try:
        report = audit_workbook(args.source)
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        raise SystemExit(f"Validación cancelada: {error}") from error
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"issueCount": report["issueCount"], "warningCount": report["warningCount"], "output": str(args.output)}, ensure_ascii=False))
    if report["issueCount"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
